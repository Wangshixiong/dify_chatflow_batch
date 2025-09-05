#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
示例测试用例生成器

这个脚本用于生成示例的测试用例Excel文件，帮助用户理解输入文件的格式。

作者: AI Assistant
创建时间: 2024
"""

import pandas as pd
from datetime import datetime
import os


def create_sample_test_cases():
    """
    创建示例测试用例
    
    Returns:
        DataFrame: 包含示例测试用例的数据框
    """
    
    # 示例测试用例数据
    test_cases = [
        # 单轮对话示例
        {
            'conversation_id': 'conv_001',
            'round': 1,
            'question': '你好，请介绍一下你的功能',
            'expected_answer': '应该包含功能介绍的回答',
            'inputs': '{}',
            'description': '基础问候和功能介绍测试'
        },
        
        # 多轮对话示例 - 保险咨询场景
        {
            'conversation_id': 'conv_002',
            'round': 1,
            'question': '我想了解车险产品',
            'expected_answer': '应该提供车险产品的基本信息',
            'inputs': '{"user_type": "个人客户"}',
            'description': '车险产品咨询 - 第1轮'
        },
        {
            'conversation_id': 'conv_002',
            'round': 2,
            'question': '车险的保费是怎么计算的？',
            'expected_answer': '应该详细说明车险保费的计算方法',
            'inputs': '{}',
            'description': '车险产品咨询 - 第2轮'
        },
        {
            'conversation_id': 'conv_002',
            'round': 3,
            'question': '我的车是2020年的本田雅阁，大概需要多少保费？',
            'expected_answer': '应该根据车型和年份给出大概的保费范围',
            'inputs': '{"car_brand": "本田", "car_model": "雅阁", "car_year": "2020"}',
            'description': '车险产品咨询 - 第3轮'
        },
        
        # 理赔咨询场景
        {
            'conversation_id': 'conv_003',
            'round': 1,
            'question': '我的车发生了交通事故，需要理赔，请问流程是什么？',
            'expected_answer': '应该提供详细的理赔流程说明',
            'inputs': '{"accident_type": "交通事故"}',
            'description': '理赔流程咨询 - 第1轮'
        },
        {
            'conversation_id': 'conv_003',
            'round': 2,
            'question': '理赔需要准备哪些材料？',
            'expected_answer': '应该列出所需的理赔材料清单',
            'inputs': '{}',
            'description': '理赔流程咨询 - 第2轮'
        },
        
        # 产品对比场景
        {
            'conversation_id': 'conv_004',
            'round': 1,
            'question': '请对比一下交强险和商业险的区别',
            'expected_answer': '应该详细对比两种保险的区别',
            'inputs': '{}',
            'description': '保险产品对比咨询'
        },
        
        # 复杂查询场景
        {
            'conversation_id': 'conv_005',
            'round': 1,
            'question': '我是新手司机，刚买了车，想要全面的保险保障，请推荐合适的保险组合',
            'expected_answer': '应该根据新手司机的特点推荐合适的保险组合',
            'inputs': '{"driver_experience": "新手", "car_status": "新车"}',
            'description': '个性化保险推荐咨询'
        },
        
        # 政策咨询场景
        {
            'conversation_id': 'conv_006',
            'round': 1,
            'question': '最新的车险改革政策有哪些变化？',
            'expected_answer': '应该介绍最新的车险改革政策内容',
            'inputs': '{}',
            'description': '政策变化咨询'
        },
        
        # 续保场景
        {
            'conversation_id': 'conv_007',
            'round': 1,
            'question': '我的车险快到期了，续保有什么优惠吗？',
            'expected_answer': '应该介绍续保的优惠政策和注意事项',
            'inputs': '{"policy_status": "即将到期"}',
            'description': '续保咨询'
        },
        
        # 异常情况测试
        {
            'conversation_id': 'conv_008',
            'round': 1,
            'question': '',  # 空问题测试
            'expected_answer': '应该提示用户输入有效问题',
            'inputs': '{}',
            'description': '空问题异常测试'
        },
        
        # 长文本测试
        {
            'conversation_id': 'conv_009',
            'round': 1,
            'question': '我想详细了解车险的各种险种，包括交强险、第三者责任险、车损险、盗抢险、玻璃单独破碎险、自燃损失险、涉水行驶险、无过失责任险、车载货物掉落责任险、车上人员责任险、不计免赔特约险等等，请分别介绍一下这些险种的保障范围、适用场景、保费计算方式以及理赔注意事项',
            'expected_answer': '应该详细介绍各种车险险种的相关信息',
            'inputs': '{}',
            'description': '长文本复杂查询测试'
        }
    ]
    
    return pd.DataFrame(test_cases)


def save_sample_file(df, filename='sample_test_cases.xlsx'):
    """
    保存示例文件
    
    Args:
        df: 数据框
        filename: 文件名
    """
    # 确保examples目录存在
    examples_dir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(examples_dir, filename)
    
    # 创建Excel写入器
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # 写入测试用例
        df.to_excel(writer, sheet_name='测试用例', index=False)
        
        # 获取工作表
        worksheet = writer.sheets['测试用例']
        
        # 调整列宽
        column_widths = {
            'A': 15,  # conversation_id
            'B': 8,   # round
            'C': 50,  # question
            'D': 40,  # expected_answer
            'E': 30,  # inputs
            'F': 25   # description
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置数据行样式
        data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
    
    print(f"示例测试用例文件已保存到: {filepath}")
    return filepath


def create_template_file():
    """
    创建空的模板文件
    """
    template_data = {
        'conversation_id': ['请填写对话ID（如：conv_001）'],
        'round': [1],
        'question': ['请填写测试问题'],
        'expected_answer': ['请填写期望的回答（可选）'],
        'inputs': ['请填写额外输入参数的JSON格式（可选，如：{"key": "value"}）'],
        'description': ['请填写测试用例描述（可选）']
    }
    
    df = pd.DataFrame(template_data)
    
    examples_dir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(examples_dir, 'test_cases_template.xlsx')
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='测试用例模板', index=False)
        
        worksheet = writer.sheets['测试用例模板']
        
        # 调整列宽
        column_widths = {
            'A': 20,  # conversation_id
            'B': 8,   # round
            'C': 50,  # question
            'D': 40,  # expected_answer
            'E': 40,  # inputs
            'F': 25   # description
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
        
        # 设置样式
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置数据行样式
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        template_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                cell.fill = template_fill
    
    print(f"测试用例模板文件已保存到: {filepath}")
    return filepath


def create_readme_for_examples():
    """
    创建示例文件的说明文档
    """
    readme_content = """
# 测试用例示例文件说明

本目录包含了 Chatflow 自动化测试工具的示例文件和模板。

## 文件说明

### 1. sample_test_cases.xlsx
这是一个包含多种测试场景的示例文件，展示了如何编写测试用例：

- **单轮对话测试**: 测试基本的问答功能
- **多轮对话测试**: 测试上下文相关的连续对话
- **业务场景测试**: 包含保险咨询、理赔、产品对比等实际业务场景
- **异常情况测试**: 测试空问题、长文本等边界情况

### 2. test_cases_template.xlsx
这是一个空的模板文件，您可以基于此模板创建自己的测试用例。

## Excel 文件格式说明

测试用例Excel文件必须包含以下列：

| 列名 | 类型 | 必需 | 说明 |
|------|------|------|------|
| conversation_id | 文本 | 是 | 对话标识符，用于区分不同的对话会话 |
| round | 数字 | 是 | 对话轮次，从1开始递增 |
| question | 文本 | 是 | 要发送给AI的问题 |
| expected_answer | 文本 | 否 | 期望的回答（用于人工对比，工具不会自动验证） |
| inputs | 文本 | 否 | 额外的输入参数，JSON格式字符串 |
| description | 文本 | 否 | 测试用例的描述信息 |

## 使用方法

1. **使用示例文件**:
   ```bash
   python main.py examples/sample_test_cases.xlsx
   ```

2. **使用模板创建自己的测试用例**:
   - 复制 `test_cases_template.xlsx`
   - 根据您的需求填写测试用例
   - 运行测试工具

## 注意事项

1. **对话ID和轮次**: 
   - 相同 conversation_id 的测试用例会被视为同一个对话
   - round 必须从1开始，并且连续递增
   - 工具会自动管理 conversation_id 的传递

2. **输入参数格式**:
   - inputs 列必须是有效的JSON格式字符串
   - 如果不需要额外参数，可以填写 `{}` 或留空

3. **问题内容**:
   - question 列不能为空
   - 支持长文本问题
   - 建议问题描述清晰、具体

4. **文件编码**:
   - 请确保Excel文件使用UTF-8编码
   - 避免使用特殊字符可能导致的编码问题

## 测试场景建议

建议您的测试用例覆盖以下场景：

1. **功能性测试**:
   - 基本问答功能
   - 业务知识查询
   - 复杂逻辑推理

2. **对话连续性测试**:
   - 多轮对话的上下文理解
   - 信息的累积和引用
   - 话题的转换和回归

3. **边界条件测试**:
   - 极长或极短的问题
   - 包含特殊字符的问题
   - 模糊或歧义的问题

4. **业务场景测试**:
   - 典型的用户咨询场景
   - 异常情况的处理
   - 不同用户类型的需求

## 结果分析

测试完成后，您将得到一个结果Excel文件，包含：
- 所有输入的测试用例信息
- AI的实际回答
- API响应时间
- 执行状态和错误信息
- 时间戳等元数据

您可以通过对比 expected_answer 和 actual_answer 来评估AI的表现。
"""
    
    examples_dir = os.path.dirname(os.path.abspath(__file__))
    readme_path = os.path.join(examples_dir, 'README.md')
    
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write(readme_content)
    
    print(f"示例文件说明文档已保存到: {readme_path}")
    return readme_path


if __name__ == '__main__':
    print("正在生成示例文件...")
    
    # 创建示例测试用例
    df = create_sample_test_cases()
    sample_file = save_sample_file(df)
    
    # 创建模板文件
    template_file = create_template_file()
    
    # 创建说明文档
    readme_file = create_readme_for_examples()
    
    print("\n示例文件生成完成！")
    print(f"- 示例测试用例: {sample_file}")
    print(f"- 测试用例模板: {template_file}")
    print(f"- 说明文档: {readme_file}")
    print("\n您可以使用这些文件来了解工具的使用方法。")